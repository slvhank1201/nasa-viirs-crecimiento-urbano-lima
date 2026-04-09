"""
    PROYECTO NASA: PREDICTOR DE CRECIMIENTO URBANO
     Análisis de Luces Nocturnas (NASA Black Marble / VIIRS)
     Aplicado a Lima Metropolitana - Perú
"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from matplotlib.gridspec import GridSpec
import warnings
warnings.filterwarnings("ignore")

LIMA_BBOX = {
    "lon_min": -77.20, "lon_max": -76.80,
    "lat_min": -12.20, "lat_max": -11.85
}
AÑOS = [2015, 2017, 2019, 2021, 2023]
GRID_SIZE = (120, 100)
DISTRITOS_LIMA = [
    ("Comas",           -11.938, -77.055),
    ("Los Olivos",      -11.993, -77.068),
    ("SMP",             -12.029, -77.063),
    ("Callao",          -12.055, -77.130),
    ("Lima Centro",     -12.045, -77.030),
    ("Miraflores",      -12.119, -77.028),
    ("San Isidro",      -12.097, -77.036),
    ("La Molina",       -12.076, -76.944),
    ("SJL",             -11.975, -76.993),
    ("VES",             -12.160, -76.970),
    ("Chorrillos",      -12.163, -77.017),
    ("Ate",             -12.026, -76.930),
]


def coord_a_pixel(lat, lon, banda_shape, bbox):
    rows, cols = banda_shape
    fila = int((bbox["lat_max"] - lat) / (bbox["lat_max"] - bbox["lat_min"]) * rows)
    col  = int((lon - bbox["lon_min"]) / (bbox["lon_max"] - bbox["lon_min"]) * cols)
    return fila, col


def agregar_distritos(ax, banda_shape, bbox):
    for nombre, lat, lon in DISTRITOS_LIMA:
        if (bbox["lat_min"] <= lat <= bbox["lat_max"] and
                bbox["lon_min"] <= lon <= bbox["lon_max"]):
            fila, col = coord_a_pixel(lat, lon, banda_shape, bbox)
            ax.plot(col, fila, "o", color="#00d4ff", markersize=4,
                    markeredgecolor="white", markeredgewidth=0.5, zorder=5)
            ax.annotate(
                nombre,
                xy=(col, fila),
                xytext=(col + 2, fila - 2),
                fontsize=5.5,
                color="white",
                fontweight="bold",
                zorder=6,
                bbox=dict(boxstyle="round,pad=0.15", facecolor="#0d1117",
                          alpha=0.65, edgecolor="none"),
            )

def load_nasa_data(usuario: str, contraseña: str, año: int) -> np.ndarray:
    import earthaccess
    import h5py
    import os

    earthaccess.login(strategy="environment")

    resultados = earthaccess.search_data(
        short_name="VNP46A3",
        temporal=(f"{año}-01", f"{año}-12"),
        bounding_box=(
            LIMA_BBOX["lon_min"], LIMA_BBOX["lat_min"],
            LIMA_BBOX["lon_max"], LIMA_BBOX["lat_max"]
        ),
    )
    print(f"{len(resultados)} archivos encontrados para {año}")

    os.makedirs("./nasa_data/", exist_ok=True)
    archivos = earthaccess.download(resultados[:1], local_path="./nasa_data/")
    archivo = archivos[0]
    print(f"Archivo descargado: {archivo}")

    with h5py.File(archivo, "r") as f:
        posibles_keys = [
            "HDFEOS/GRIDS/VIIRS_Grid_DNB_2d/Data Fields/AllAngle_Composite_Snow_Free",
            "HDFEOS/GRIDS/VIIRS_Grid_DNB_2d/Data Fields/NearNadir_Composite_Snow_Free",
        ]

        banda = None
        for key in posibles_keys:
            if key in f:
                raw   = f[key][:]
                banda = raw.astype(np.float32)
                scale = float(f[key].attrs.get("scale_factor", 0.1))
                banda[banda >= 65534] = np.nan
                banda = banda * scale
                banda[banda <= 0] = np.nan
                lat = f["HDFEOS/GRIDS/VIIRS_Grid_DNB_2d/Data Fields/lat"][:]
                lon = f["HDFEOS/GRIDS/VIIRS_Grid_DNB_2d/Data Fields/lon"][:]

                idx_lat = np.where((lat >= LIMA_BBOX["lat_min"]) & (lat <= LIMA_BBOX["lat_max"]))[0]
                idx_lon = np.where((lon >= LIMA_BBOX["lon_min"]) & (lon <= LIMA_BBOX["lon_max"]))[0]

                if len(idx_lat) > 0 and len(idx_lon) > 0:
                    banda = banda[np.ix_(idx_lat, idx_lon)]
                    print(f"   Recortado a Lima: {banda.shape}")
                break

        if banda is None:
            raise KeyError("No se encontró el dataset esperado.")
    return banda


def calcular_delta_luminosidad(datos):
    años_lista = sorted(datos.keys())
    deltas = {}
    for i in range(1, len(años_lista)):
        año_ant = años_lista[i - 1]
        año_act = años_lista[i]
        deltas[f"{año_ant}→{año_act}"] = datos[año_act] - datos[año_ant]
    return deltas


def detectar_zonas_expansion(datos, umbral_percentil=85):
    primer_año = min(datos.keys())
    último_año = max(datos.keys())
    delta_total = datos[último_año] - datos[primer_año]

    umbral = np.percentile(delta_total[delta_total > 0], umbral_percentil)
    zonas_expansion = (delta_total > umbral).astype(np.uint8)

    print(f"\nZonas de alta expansión detectadas: {zonas_expansion.sum()} píxeles")
    print(f"   Umbral de radiancia: {umbral:.2f} nW/cm²/sr")
    print(f"   Crecimiento total promedio: {np.nanmean(delta_total):.2f}")

    return zonas_expansion, delta_total


def calcular_estadisticas(datos):
    umbral_urbano = 10
    registros = []
    for año, grilla in sorted(datos.items()):
        grilla_valida = np.nan_to_num(grilla, nan=0.0)
        registros.append({
            "Año": año,
            "Luz_Total_GW": np.nansum(grilla) / 1e6,
            "Area_Iluminada_km2": (grilla_valida > umbral_urbano).sum() * 0.25,
            "Radiancia_Media": np.nanmean(grilla),
            "Radiancia_Max": np.nanmax(grilla),
            "Cobertura_Pct": (grilla_valida > umbral_urbano).mean() * 100,
        })

    df = pd.DataFrame(registros)
    df["Crecimiento_Luz_Pct"] = df["Luz_Total_GW"].pct_change().mul(100).round(1).fillna(0)
    df["Luz_Total_GW"]        = df["Luz_Total_GW"].round(4)
    df["Radiancia_Media"]     = df["Radiancia_Media"].round(2)
    df["Cobertura_Pct"]       = df["Cobertura_Pct"].round(1)
    return df

def predecir_crecimiento(df_stats, año_objetivo=2026):
    from sklearn.linear_model import LinearRegression

    X = df_stats[["Año"]].values
    y = df_stats["Luz_Total_GW"].values

    modelo = LinearRegression()
    modelo.fit(X, y)

    pred_val = modelo.predict([[año_objetivo]])[0]
    r2       = modelo.score(X, y)

    print(f"\nPREDICCIÓN {año_objetivo}:")
    print(f"   Luz total estimada: {pred_val:.2f} GW·equivalente")
    print(f"   R²: {r2:.4f}")
    print(f"   Crecimiento anual: {modelo.coef_[0]:.4f} GW/año")

    return {"modelo": modelo, "pred_2026": pred_val, "r2": r2, "coef_anual": modelo.coef_[0]}


def proyectar_mapas(datos, años_futuros):
    años_hist = np.array(sorted(datos.keys())).reshape(-1, 1)
    primer_año = min(datos.keys())
    rows, cols = datos[primer_año].shape

    stack   = np.stack([datos[a] for a in sorted(datos.keys())], axis=0)
    pixeles = stack.reshape(len(años_hist), -1).astype(np.float64)

    X      = años_hist.flatten().astype(np.float64)
    X_mean = X.mean()
    X_var  = ((X - X_mean) ** 2).sum()

    pendiente  = ((X - X_mean).reshape(-1, 1) * pixeles).sum(axis=0) / X_var
    intercepto = pixeles.mean(axis=0) - pendiente * X_mean

    print(f"\n[PROYECCION] Pendiente promedio: {pendiente.mean():.4f} nW/cm²/sr por año")

    mapas_proyectados = {}
    for año in años_futuros:
        mapa = (intercepto + pendiente * año).reshape(rows, cols).astype(np.float32)
        mapa = np.clip(mapa, 0.0, 300.0)
        mapa[np.isnan(datos[primer_año])] = np.nan
        mapas_proyectados[año] = mapa

        pct = ((np.nanmean(mapa) - np.nanmean(datos[max(datos.keys())])) /
                np.nanmean(datos[max(datos.keys())])) * 100
        print(f"   {año}: radiancia media = {np.nanmean(mapa):.2f}  ({pct:+.1f}% vs {max(datos.keys())})")

    return mapas_proyectados


def visualizar_dashboard(datos, zonas, delta_total, df_stats, pred, mapas_proyectados=None):
    import numpy.ma as ma
    import matplotlib.ticker as mticker

    año_ini = min(datos.keys())
    año_fin = max(datos.keys())
    banda_shape = datos[año_fin].shape
    n_rows, n_cols = banda_shape

    años_proy_lista = sorted(mapas_proyectados.keys()) if mapas_proyectados else []
    n_filas = 3 if años_proy_lista else 2

    lat_vals  = np.linspace(LIMA_BBOX["lat_max"], LIMA_BBOX["lat_min"], 5)
    lon_vals  = np.linspace(LIMA_BBOX["lon_min"], LIMA_BBOX["lon_max"], 5)
    row_ticks = np.linspace(0, n_rows - 1, 5)
    col_ticks = np.linspace(0, n_cols - 1, 5)

    fig = plt.figure(figsize=(22, 10 + (4 if n_filas == 3 else 0)), facecolor="#0d1117")
    fig.suptitle("ANALISIS DE CRECIMIENTO URBANO  ·  NASA Black Marble  ·  Lima, Peru",
                 fontsize=15, fontweight="bold", color="white", y=0.995)

    gs = GridSpec(n_filas, max(3, len(años_proy_lista)), figure=fig, hspace=0.55, wspace=0.30)

    todos_mapas  = list(datos.values()) + (list(mapas_proyectados.values()) if mapas_proyectados else [])
    vmax_global  = np.nanpercentile(np.concatenate([m[~np.isnan(m)].flatten() for m in todos_mapas]), 99)

    def estilizar_mapa(ax, titulo, color_titulo="white"):
        ax.set_facecolor("#0d1117")
        ax.set_xticks(col_ticks)
        ax.set_xticklabels([f"{v:.2f}" for v in lon_vals], fontsize=5.5, color="#8b949e")
        ax.set_yticks(row_ticks)
        ax.set_yticklabels([f"{v:.2f}" for v in lat_vals], fontsize=5.5, color="#8b949e")
        ax.set_xlabel("Longitud", fontsize=6, color="#8b949e")
        ax.set_ylabel("Latitud",  fontsize=6, color="#8b949e")
        ax.set_title(titulo, color=color_titulo, fontsize=10, pad=4)
        for spine in ax.spines.values():
            spine.set_color("#30363d")

    def add_colorbar(im, ax, label):
        cb = plt.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
        cb.set_label(label, color="white", fontsize=7)
        cb.ax.yaxis.set_tick_params(color="white", labelsize=6)
        plt.setp(cb.ax.yaxis.get_ticklabels(), color="white")

    ax_a = fig.add_subplot(gs[0, 0])
    add_colorbar(ax_a.imshow(datos[año_ini], cmap="inferno", vmin=0, vmax=vmax_global, aspect="auto"), ax_a, "nW/cm²/sr")
    agregar_distritos(ax_a, banda_shape, LIMA_BBOX)
    estilizar_mapa(ax_a, f"[REAL] Luces Nocturnas {año_ini}")

    ax_b = fig.add_subplot(gs[0, 1])
    add_colorbar(ax_b.imshow(datos[año_fin], cmap="inferno", vmin=0, vmax=vmax_global, aspect="auto"), ax_b, "nW/cm²/sr")
    agregar_distritos(ax_b, banda_shape, LIMA_BBOX)
    estilizar_mapa(ax_b, f"[REAL] Luces Nocturnas {año_fin}")

    ax_c = fig.add_subplot(gs[0, 2])
    delta_plot = np.nan_to_num(delta_total.copy())
    lim = np.percentile(np.abs(delta_plot), 95)
    add_colorbar(ax_c.imshow(delta_plot, cmap="RdYlGn", vmin=-lim, vmax=lim, aspect="auto"), ax_c, "Delta nW/cm²/sr")
    agregar_distritos(ax_c, banda_shape, LIMA_BBOX)
    estilizar_mapa(ax_c, f"Delta Radiancia {año_ini}→{año_fin}\n(verde = crecimiento)")

    ax_d = fig.add_subplot(gs[1, 0])
    ax_d.imshow(datos[año_fin], cmap="gray", alpha=0.55, aspect="auto")
    ax_d.imshow(ma.masked_where(zonas == 0, zonas), cmap="autumn", alpha=0.85, aspect="auto")
    agregar_distritos(ax_d, banda_shape, LIMA_BBOX)
    estilizar_mapa(ax_d, "Zonas de Expansion Urbana\n(candidatas a nueva infraestructura)")

    ax_e = fig.add_subplot(gs[1, 1])
    ax_e.set_facecolor("#161b22")
    ax_e.plot(df_stats["Año"].values, df_stats["Luz_Total_GW"].values,
              "o-", color="#58a6ff", linewidth=2.5, markersize=8, label="Datos NASA reales", zorder=3)
    años_proy_arr = np.array([año_fin, 2024, 2025, 2026, 2027, 2028, 2029, 2030])
    luz_proy_arr  = pred["modelo"].predict(años_proy_arr.reshape(-1, 1))
    ax_e.fill_between(años_proy_arr, luz_proy_arr * 0.90, luz_proy_arr * 1.10,
                      color="#f78166", alpha=0.15, label="Intervalo ±10%")
    ax_e.plot(años_proy_arr, luz_proy_arr, "--o", color="#f78166",
              linewidth=2.2, markersize=5, label="Proyeccion lineal", zorder=3)
    for a, v in zip(años_proy_arr[1:], luz_proy_arr[1:]):
        ax_e.annotate(f"{v:.3f}", xy=(a, v), xytext=(0, 9), textcoords="offset points",
                      ha="center", fontsize=7, color="#f78166")
    ax_e.axvline(x=año_fin + 0.3, color="#8b949e", linestyle=":", alpha=0.7)
    ax_e.set_title("Tendencia de Luminosidad Urbana + Proyeccion 2030", color="white", fontsize=10)
    ax_e.set_xlabel("Año", color="white", fontsize=9)
    ax_e.set_ylabel("Luz Total (GW equiv.)", color="white", fontsize=9)
    ax_e.tick_params(colors="white", axis="x", rotation=45)
    ax_e.xaxis.set_major_locator(mticker.MultipleLocator(2))
    ax_e.legend(facecolor="#21262d", labelcolor="white", fontsize=8)
    ax_e.spines[:].set_color("#30363d")
    ax_e.grid(color="#21262d", linestyle="--", alpha=0.5)
    ax_e.set_xlim(df_stats["Año"].values[0] - 0.5, 2030.5)

    ax_f = fig.add_subplot(gs[1, 2])
    ax_f.axis("off")
    tabla_data = df_stats[["Año", "Luz_Total_GW", "Cobertura_Pct", "Crecimiento_Luz_Pct"]].copy()
    tabla_data.columns = ["Año", "Luz (GW)", "Cob. %", "Delta %"]
    años_extra = [2026, 2027, 2028, 2029, 2030]
    luz_extra  = pred["modelo"].predict(np.array(años_extra).reshape(-1, 1))
    for a, v in zip(años_extra, luz_extra):
        tabla_data = tabla_data._append(
            {"Año": a, "Luz (GW)": round(v, 4), "Cob. %": "-", "Delta %": "(proy.)"},
            ignore_index=True)
    tabla = ax_f.table(cellText=tabla_data.values, colLabels=tabla_data.columns,
                       cellLoc="center", loc="center")
    tabla.auto_set_font_size(False)
    tabla.set_fontsize(9)
    tabla.scale(1.1, 1.5)
    n_hist = len(df_stats)
    for (fila, col), celda in tabla.get_celld().items():
        if fila == 0:
            celda.set_facecolor("#21262d"); celda.set_text_props(color="#8b949e", fontweight="bold")
        elif fila > n_hist:
            celda.set_facecolor("#2d1f1a"); celda.set_text_props(color="#f78166")
        else:
            celda.set_facecolor("#161b22"); celda.set_text_props(color="white")
        celda.set_edgecolor("#30363d")
    ax_f.set_title("Resumen KPIs + Proyeccion 2026-2030", color="white", fontsize=10)

    if años_proy_lista:
        for i, año in enumerate(años_proy_lista):
            ax_p = fig.add_subplot(gs[2, i])
            mapa  = mapas_proyectados[año]
            media = np.nanmean(mapa)
            pct   = ((media - np.nanmean(datos[año_fin])) / np.nanmean(datos[año_fin])) * 100
            add_colorbar(ax_p.imshow(mapa, cmap="magma", vmin=0, vmax=vmax_global, aspect="auto"), ax_p, "nW/cm²/sr")
            agregar_distritos(ax_p, banda_shape, LIMA_BBOX)
            estilizar_mapa(ax_p, f"[PROYECCION] {año}\nmedia={media:.1f}  ({pct:+.1f}% vs {año_fin})",
                           color_titulo="#f78166")

    plt.savefig("nasa_urban_growth_dashboard.png", dpi=150, bbox_inches="tight", facecolor="#0d1117")
    print("\n[OK] Dashboard guardado: nasa_urban_growth_dashboard.png")
    plt.show()


def exportar_csv(datos, zonas, delta_total, mapas_proyectados=None):
    año_ini = min(datos.keys())
    año_fin = max(datos.keys())
    n_rows, n_cols = datos[año_fin].shape

    lat_vals = np.linspace(LIMA_BBOX["lat_max"], LIMA_BBOX["lat_min"], n_rows)
    lon_vals = np.linspace(LIMA_BBOX["lon_min"], LIMA_BBOX["lon_max"], n_cols)

    n_años  = año_fin - año_ini
    val_ini = np.nan_to_num(datos[año_ini], nan=0.001)
    val_fin = np.nan_to_num(datos[año_fin], nan=0.001)
    with np.errstate(divide="ignore", invalid="ignore"):
        cagr = np.where(val_ini > 0, (val_fin / val_ini) ** (1 / n_años) - 1, 0.0)

    registros = []
    for r in range(0, n_rows, 2):
        for c in range(0, n_cols, 2):
            rad_ini = float(np.nan_to_num(datos[año_ini][r, c]))
            rad_fin = float(np.nan_to_num(datos[año_fin][r, c]))
            tasa    = float(cagr[r, c])

            if tasa > 0.05:   tendencia = "crecimiento_alto"
            elif tasa > 0.01: tendencia = "crecimiento_moderado"
            elif tasa > -0.01: tendencia = "estable"
            else:              tendencia = "decrecimiento"

            fila = {
                "latitud":         round(lat_vals[r], 5),
                "longitud":        round(lon_vals[c], 5),
                f"rad_{año_ini}":  round(rad_ini, 2),
                f"rad_{año_fin}":  round(rad_fin, 2),
                "delta_historico": round(float(np.nan_to_num(delta_total[r, c])), 2),
                "cagr_pct":        round(tasa * 100, 3),
                "tendencia":       tendencia,
                "zona_expansion":  int(zonas[r, c]),
            }
            if mapas_proyectados:
                for año_p, mapa_p in sorted(mapas_proyectados.items()):
                    val_p = float(np.nan_to_num(mapa_p[r, c]))
                    fila[f"rad_{año_p}"]              = round(val_p, 2)
                    fila[f"delta_vs_{año_fin}_{año_p}"] = round(val_p - rad_fin, 2)
            registros.append(fila)

    df_export = pd.DataFrame(registros)
    df_export["nivel_urbano"] = pd.cut(
        df_export[f"rad_{año_fin}"],
        bins=[0, 10, 30, 60, 120, 999],
        labels=["rural", "peri-urbano", "urbano", "urbano_denso", "centro_comercial"]
    )
    df_export.to_csv("nasa_datos_lima.csv", index=False)
    print(f"[OK] CSV exportado: nasa_datos_lima.csv ({len(df_export):,} filas x {len(df_export.columns)} cols)")
    return df_export


def exportar_powerbi_excel(datos, zonas, df_stats, pred, mapas_proyectados=None,
                           ruta="nasa_lima_powerbi.xlsx"):
 
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HDR_FILL  = PatternFill('solid', start_color='1F3864')   
    REAL_FILL = PatternFill('solid', start_color='D9E1F2')  
    PROY_FILL = PatternFill('solid', start_color='FCE4D6') 
    HDR_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    BODY_FONT = Font(name='Arial', size=10)
    CENTER    = Alignment(horizontal='center', vertical='center')
    thin      = Side(style='thin', color='BFBFBF')
    BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_sheet(ws, df_data, title):
        ws.title = title
        cols = list(df_data.columns)
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.font = HDR_FONT; c.fill = HDR_FILL
            c.alignment = CENTER; c.border = BORDER
            ws.column_dimensions[get_column_letter(ci)].width = max(20, len(col) + 4)
        for ri, row in df_data.iterrows():
            fill = REAL_FILL if str(row.get('tipo', '')).startswith('R') else PROY_FILL
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri + 2, column=ci, value=val)
                c.font = BODY_FONT; c.fill = fill
                c.alignment = CENTER; c.border = BORDER
        ws.row_dimensions[1].height = 22

    año_fin      = max(datos.keys())
    años_hist    = sorted(datos.keys())
    años_proy    = sorted(mapas_proyectados.keys()) if mapas_proyectados else []
    todos_años   = años_hist + años_proy
    umbral       = 10
    bins         = [0, 10, 30, 60, 120, 9999]
    niveles      = ['rural', 'peri-urbano', 'urbano', 'urbano_denso', 'centro_comercial']

    def get_vals(año):
        if año in datos:
            return datos[año].flatten()
        return mapas_proyectados[año].flatten()

    kpi_rows = []
    for año in todos_años:
        vals  = np.nan_to_num(get_vals(año), nan=0.0)
        media = round(float(np.nanmean(vals)), 2)
        cob   = round(float((vals > umbral).mean() * 100), 1)
        tipo  = 'Real' if año <= año_fin else 'Proyección'
        zona  = int(zonas.sum()) if año == año_fin else ''
        kpi_rows.append({
            'anio':                    año,
            'tipo':                    tipo,
            'radiancia_media_nW':      media,
            'cobertura_iluminada_pct': cob,
            'pixeles_zona_expansion':  zona,
        })
    df_kpi = pd.DataFrame(kpi_rows)

    tend_rows = []
    for año in todos_años:
        vals  = np.nan_to_num(get_vals(año), nan=0.0)
        media = round(float(np.nanmean(vals)), 2)
        tipo  = 'Real' if año <= año_fin else 'Proyección'
        tend_rows.append({'anio': año, 'tipo': tipo, 'radiancia_media_nW': media})
    df_tend = pd.DataFrame(tend_rows)

    dist_rows = []
    for año in todos_años:
        vals = np.nan_to_num(get_vals(año), nan=0.0)
        tipo = 'Real' if año <= año_fin else 'Proyección'
        cats = pd.cut(vals, bins=bins, labels=niveles)
        counts = pd.Series(cats).value_counts()
        for nivel in niveles:
            dist_rows.append({
                'anio':             año,
                'tipo':             tipo,
                'nivel_urbano':     nivel,
                'cantidad_pixeles': int(counts.get(nivel, 0)),
            })
    df_dist = pd.DataFrame(dist_rows)

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb.create_sheet(), df_kpi,  'KPIs_por_año')
    write_sheet(wb.create_sheet(), df_tend, 'Tendencia_Radiancia')
    write_sheet(wb.create_sheet(), df_dist, 'Distribucion_Nivel_Urbano')
    wb.save(ruta)

    print(f"\n[OK] Excel Power BI exportado: {ruta}")
    print(f"     KPIs_por_año          → {len(df_kpi)} filas")
    print(f"     Tendencia_Radiancia   → {len(df_tend)} filas")
    print(f"     Distribucion_Nivel    → {len(df_dist)} filas")
    return df_kpi, df_tend, df_dist

if __name__ == "__main__":
    print("=" * 60)
    print(" NASA URBAN GROWTH ANALYZER - Lima, Perú")
    print("=" * 60)

    print("\n Cargando datos satelitales...")
    datos = {}
    for año in [2019, 2021, 2023]:
        print(f"\n   Descargando año {año}...")
        datos[año] = load_nasa_data("slv_hank", "scxLp89_@ppqt", año)
    print(f"{len(datos)} capas cargadas: {list(datos.keys())}")
    print("\n Calculando variación de luminosidad...")
    zonas, delta_total = detectar_zonas_expansion(datos, umbral_percentil=85)

    print("\n Calculando KPIs por período...")
    df_stats = calcular_estadisticas(datos)
    print(df_stats.to_string(index=False))

    print("\n Entrenando modelo de predicción...")
    pred = predecir_crecimiento(df_stats, año_objetivo=2026)

    print("\n Proyectando mapas 2026-2030...")
    mapas_proyectados = proyectar_mapas(datos, años_futuros=[2026, 2027, 2028, 2029, 2030])


    print("\n Generando dashboard...")
    visualizar_dashboard(datos, zonas, delta_total, df_stats, pred,
                         mapas_proyectados=mapas_proyectados)

  
    print("\n Exportando CSV completo...")
    df_export = exportar_csv(datos, zonas, delta_total, mapas_proyectados=mapas_proyectados)

    print("\n Exportando Excel limpio para Power BI...")
    exportar_powerbi_excel(
        datos=datos,
        zonas=zonas,
        df_stats=df_stats,
        pred=pred,
        mapas_proyectados=mapas_proyectados,
        ruta="nasa_lima_powerbi.xlsx"
    )

    print("\n" + "=" * 60)
    print("ANÁLISIS COMPLETADO")
    print(" Dashboard PNG  → nasa_urban_growth_dashboard.png")
    print(" CSV completo   → nasa_datos_lima.csv")
    print(" Excel Power BI → nasa_lima_powerbi.xlsx")
    print("=" * 60)